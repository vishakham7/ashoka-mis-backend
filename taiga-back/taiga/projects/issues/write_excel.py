from openpyxl import *
import openpyxl
import io
import os
from django.utils.translation import ugettext as _
from django.conf import settings
from django.shortcuts import render
from taiga.base.utils import db, text
from taiga.projects.issues.apps import (
    connect_issues_signals,
    disconnect_issues_signals)
from taiga.projects.votes.utils import attach_total_voters_to_queryset
from taiga.projects.notifications.utils import attach_watchers_to_queryset
from datetime import date, datetime, timedelta
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK, BORDER_MEDIUM
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl.writer.excel import save_virtual_workbook
from django.http import HttpResponse
from xhtml2pdf import pisa 
import pdfkit
import PyPDF2
import urllib3
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor, OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from taiga.users.models import User
# from PIL import Image
from bs4 import BeautifulSoup 
from taiga.projects.models import Membership



def style(ws,fieldnames, issue,file_name=None,Compliance_file_name=None):
    font = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    font2 = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FFFFFF')
    color = Font(name='Calibri',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='696969')
    border = Border(
            left=Side(border_style=BORDER_THIN, color='FF000000'),
            right=Side(border_style=BORDER_THIN, color='FF000000'),
            top=Side(border_style=BORDER_THIN, color='FF000000'),
            bottom=Side(border_style=BORDER_THIN, color='FF000000')
        )
    fill=PatternFill(start_color = '00C0C0C0',
            end_color = '00C0C0C0',
            fill_type = 'solid')
    dd = Font(underline='single', color='000000FF')
    row_count = ws.max_row
    column_count = ws.max_column
    # for cell in ws['2:2']:
    #     cell.font = font

    for cell2 in ws['9:9']:
        cell2.fill = fill
        cell2.font = font

    for cell3 in ws['10:10']:
        cell3.fill = fill
        cell3.font = font

    for i in range(3,row_count+1):
        ws.row_dimensions[i].height = 50

    ws.row_dimensions[2].height = 40
    ws.row_dimensions[1].height = 40
    
    for row in ws:
        for cell1 in row:
            cell1.border = border
            cell1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    for i0 in ws['1:1']:
        i0.alignment = Alignment(horizontal='left', vertical='center')
    for i1 in ws['2:2']:
        i1.alignment = Alignment(horizontal='left', vertical='center')

    for i2 in ws['3:3']:
        i2.alignment = Alignment(horizontal='left', vertical='center')
    for i3 in ws['4:4']:
        i3.alignment = Alignment(horizontal='left', vertical='center')
    for i4 in ws['5:5']:
        i4.alignment = Alignment(horizontal='left', vertical='center')
    for i5 in ws['6:6']:
        i5.alignment = Alignment(horizontal='left', vertical='center')
    for i6 in ws['7:7']:
        i6.alignment = Alignment(horizontal='left', vertical='center')
    for i7 in ws['8:8']:
        i7.alignment = Alignment(horizontal='left', vertical='center')


    column_widths = []
    for row in fieldnames:
        for i in range(len(row)):
            if len(column_widths) > i:
                if len(row) > column_widths[i]:
                    column_widths[i] = len(row)
            else:
                column_widths += [len(row)]
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width


    #////////////////////////// working images
    if file_name:
        file_row = []
        l = []
        for row in range(10,row_count+1):
            file_row.append(row)
        
        for i in range(len(file_row)):
            if len(file_row)==(row_count-9):
                l.append(file_row[i])
        file_name = []
        split = []
        aaa=[]
        val = []
        n=""
        hh=""
        for new_row in l:
            ws.row_dimensions[new_row].height = 140
            ws.column_dimensions[get_column_letter(6)].width = 50
            file = ws.cell(row=new_row, column=6).value
            if file:
                split = file.split('\n')
                if split:
                    aaa.append(split)
                for aa in aaa:

                    for j in range(len(aa)-1):
                        
                        new = aa[j].split('.')
                        doc_name = new[-2].split('/')
                        file_name = doc_name[-1]+'.'+new[-1]
        
                        name = ws.cell(row=new_row, column=6).value
                        n += name
                        
                        if new[-1]=="jpeg" or new[-1]=="png" or new[-1]=="jpg":
                            if len(aa) == 4:
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[0])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height = 120
                                    img.width = 120
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 1))
                                        cellw = lambda x: c2e((x *0.09))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 5
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        print(coloffset, rowoffset)
                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        # ws.cell(row=new_row, column=6).value = "<img src='"+ aa[0] + "' height=100 width=70/><br>"

                                        # ws.cell(row=new_row, column=6).font = font2
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[1])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height=200
                                    img.width =200
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 8))
                                        cellw = lambda x: c2e((x *0.01))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 5
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        print(coloffset, rowoffset)
                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        # ws.cell(row=new_row, column=6).value = "<img src='"+ aa[1] + "' height=100 width=70/><br>"

                                        # ws.cell(row=new_row, column=6).font = font2
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[0])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height=120
                                    img.width =120
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 16))
                                        cellw = lambda x: c2e((x *0.01))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 5
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        print(coloffset, rowoffset)
                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        ws.cell(row=new_row, column=6).value = "<img src='"+ aa[2] + "' height=100 width=70/><br><img src='"+ aa[1] + "' height=100 width=70/><br><img src='"+ aa[0] + "' height=100 width=70/>"

                                        ws.cell(row=new_row, column=6).font = font2
                            elif len(aa) == 3:
                                print("00000000000000000000000000")
                                print(len(aa))
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[0])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height=120
                                    img.width =120
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 1))
                                        cellw = lambda x: c2e((x *0.09))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 5
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        print(coloffset, rowoffset)
                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        # ws.cell(row=new_row, column=6).value = "<img src='"+ aa[1] + "' height=100 width=70/><br>"

                                        # ws.cell(row=new_row, column=6).font = font2
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[1])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height=120
                                    img.width =120
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 8))
                                        cellw = lambda x: c2e((x *0.01))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 5
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        print(coloffset, rowoffset)
                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        ws.cell(row=new_row, column=6).value = "<img src='"+ aa[1] + "' height=100 width=70/><br><img src='"+ aa[0] + "' height=100 width=70/><br>"

                                        ws.cell(row=new_row, column=6).font = font2
                                    
                            elif len(aa) == 2:
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[0])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height=120
                                    img.width =120
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 1))
                                        cellw = lambda x: c2e((x *0.09))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 5
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        ws.cell(row=new_row, column=6).value = "<img src='"+ aa[0] + "' height=100 width=70/><br>"

                                        ws.cell(row=new_row, column=6).font = font2

                        
                        if new[-1]=="xlsx" or new[-1]=="docx" or new[-1]=="doc" or new[-1]=="pdf" or new[-1]=="mp3":
                            ws.cell(row=new_row, column=6).hyperlink = aa[j]
                            ws.cell(row=new_row, column=6).value = file_name

def comp(ws,Compliance_file_name):
    font2 = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FFFFFF')
    if Compliance_file_name:
        row_count = ws.max_row
        column_count = ws.max_column
        file_row = []
        for row in range(10,row_count+1):
            file_row.append(row)
        l=[]
        for i in range(len(file_row)):
            if len(file_row)==(row_count-9):
                l.append(file_row[i])
        
        file_name = []
        split = []
        aaa=[]
        val = []
        n=""
        hh=""
        for new_row in l:
            ws.row_dimensions[new_row].height = 140
            ws.column_dimensions[get_column_letter(7)].width = 50
            file = ws.cell(row=new_row, column=7).value
            if file:
                split = file.split('\n')
                if split:
                    aaa.append(split)
                for aa in aaa:

                    for j in range(len(aa)-1):
                        
                        new = aa[j].split('.')
                        doc_name = new[-2].split('/')
                        file_name = doc_name[-1]+'.'+new[-1]
        
                        name = ws.cell(row=new_row, column=7).value
                        n += name

                        if new[-1]=="jpeg" or new[-1]=="png" or new[-1]=="jpg":
                            
                            if len(aa) == 4:
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[0])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height=120
                                    img.width =120
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 1))
                                        cellw = lambda x: c2e((x *0.09))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 6
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        print(coloffset, rowoffset)
                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        # ws.cell(row=new_row, column=6).value = "<img src='"+ aa[0] + "' height=100 width=70/><br>"

                                        # ws.cell(row=new_row, column=6).font = font2
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[1])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height=120
                                    img.width =120
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 8))
                                        cellw = lambda x: c2e((x *0.01))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 6
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        print(coloffset, rowoffset)
                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        # ws.cell(row=new_row, column=6).value = "<img src='"+ aa[1] + "' height=100 width=70/><br>"

                                        # ws.cell(row=new_row, column=6).font = font2
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[0])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height=120
                                    img.width =120
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 16))
                                        cellw = lambda x: c2e((x *0.01))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 6
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        print(coloffset, rowoffset)
                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        ws.cell(row=new_row, column=6).value = "<img src='"+ aa[2] + "' height=100 width=70/><br><img src='"+ aa[1] + "' height=100 width=70/><br><img src='"+ aa[0] + "' height=100 width=70/>"

                                        ws.cell(row=new_row, column=6).font = font2
                            elif len(aa) == 3:
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[0])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height=120
                                    img.width =120
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 1))
                                        cellw = lambda x: c2e((x *0.09))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 6
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        print(coloffset, rowoffset)
                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        # ws.cell(row=new_row, column=6).value = "<img src='"+ aa[1] + "' height=100 width=70/><br>"

                                        # ws.cell(row=new_row, column=6).font = font2
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[1])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height=120
                                    img.width =120
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 8))
                                        cellw = lambda x: c2e((x *0.01))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 6
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        print(coloffset, rowoffset)
                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        ws.cell(row=new_row, column=7).value = "<img src='"+ aa[1] + "' height=100 width=70/><br><img src='"+ aa[0] + "' height=100 width=70/><br>"

                                        ws.cell(row=new_row, column=7).font = font2
                                    
                            elif len(aa) == 2:
                                http = urllib3.PoolManager()
                                r = http.request('GET',aa[0])
                                image_file = io.BytesIO(r.data)
                                
                                if image_file:

                                    img = Image(image_file)

                                    img.height=120
                                    img.width =120
                                   
                                    if img:
                                        c2e = cm_to_EMU
                                        p2e = pixels_to_EMU

                                        h, w = img.height, img.width
                                        # Calculated number of cells width or height from cm into EMUs
                                        # celh = [1,8,16]
                                        # celw = [0.09,0.01,0.01]
                                        # cellh = 0
                                        # cellw = 0
                                        cellh = lambda x: c2e((x * 1))
                                        cellw = lambda x: c2e((x *0.09))
                                        # cellw = cellzw
                                        # print(cellw, cellh)
                                        # Want to place image in row 5 (6 in excel), column 2 (C in excel)
                                        # Also offset by half a column.
                                        column = 6
                                        # colof = [28000]
                                        coloffset = cellh(0.5)
                                        # coloffset = 2880000
                                        row = new_row-1
                                        rowoffset = cellw(0.03)
                                        rowpp = [107,197]
                                

                                        print(coloffset, rowoffset)
                                        size = XDRPositiveSize2D(p2e(h), p2e(w))
                                        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
                                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                                        
                                        ws.add_image(img)
                                        
                                        ws.cell(row=new_row, column=7).value = "<img src='"+ aa[0] + "' height=100 width=70/><br>"

                                        ws.cell(row=new_row, column=7).font = font2

                                
                        if new[-1]=="xlsx" or new[-1]=="docx" or new[-1]=="doc" or new[-1]=="pdf" or new[-1]=="mp3":
                            ws.cell(row=new_row, column=7).hyperlink = aa[j]
                            ws.cell(row=new_row, column=7).value = file_name
                        
                                    
                        # # if new[-1]=="svg":
                        # #     http = urllib3.PoolManager()
                        # #     # r = http.request('GET', aa[j-(len(aa)-1)])
                        # #     if aa[j]:
                        # #         r = http.request('GET', aa[j])
                        # #         image_file = io.BytesIO(r.data)
                        # #         if image_file:
                        # #             img = Image(image_file)
                        # #             img.height=100
                        # #             img.width =100
                        # #             img.format = new[-1]

                        # #             ws.add_image(img,'F'+str(new_row))
                        # #             ws.row_dimensions[new_row].height = 40
                        # #             # ws.cell(row=new_row, column=7).value = aa[j]
                        # #             # ws.cell(row=new_row, column=7).hyperlink = aa[j]
                        # #             ws.cell(row=new_row, column=6).value = "<img src='"+ aa[j] + "' height=100 width=70/>"

                        # #             ws.cell(row=new_row, column=6).font = font2
                        


                        #     # print(aa[j])
                        #     # print(aa[j-(len(aa))])
                        #     # http = urllib3.PoolManager()
                        #     # # r = http.request('GET', aa[j-(len(aa)-1)])
                        #     # r = http.request('GET', aa[j])
                        #     # image_file = io.BytesIO(r.data)
                        
                        #     # img = Image(image_file)
                        #     # img.height=100
                        #     # img.width =100
                        #     # ws.add_image(img,'G'+str(new_row))
                        #     # # ws.cell(row=new_row, column=7).value = aa[j]
                        #     # # ws.cell(row=new_row, column=7).hyperlink = aa[j]
                        #     # ws.cell(row=new_row, column=7).value = "<img src='"+  aa[j] + "' height=100 width=70/>"
                        # if len(n)>180:                
                        #     ws.row_dimensions[new_row].height = 120
                        #     if new[-1]=="jpeg" or new[-1]=="jpg" or new[-1]=="png":
                        # #         # print(aa[j])
                        #         http = urllib3.PoolManager()
                        #         # r = http.request('GET', aa[j-(len(aa)-1)])
                        #         r = http.request('GET', aa[j])
                        # # #         print(aa[j])
                        #         image_file = io.BytesIO(r.data)
                        # # #         print("0000---------------0000")
                        # # #         print(image_file)
                        #         img = Image(image_file)
                        # # #         print("000000000000000000000000")
                        # # #         print(img)
                        #         img.height=100
                        #         img.width =100
                        #         if img:
                        #             ws.add_image(img,'F'+str(new_row))
                        #     #         # ws.cell(row=new_row, column=7).value = aa[j]
                        #             ws.cell(row=new_row, column=6).value = "<img scr='"+  aa[j] + "'  />"
                        #             ws.cell(row=new_row, column=6).font = font2

                                

                        #         # ============================================================
                        #         # r1 = http.request('GET', aa[0])
                        #         # # r = http.request('GET', aa[j])
                        #         # image_file1 = io.BytesIO(r1.data)
                            
                        #         # img1 = Image(image_file1)
                        #         # img1.height=100
                        #         # img1.width =100
                        #         # ws.add_image(img1,'G'+str(new_row))
                                
                        #         # ws.cell(row=new_row, column=7).value = "<img scr='"+  aa[j-(len(aa)-1)] + "'></img>"
                        #         # ws.cell(row=new_row, column=7).value = '<img src="' + aa[0] + '"/>'
                        #         ws.cell(row=new_row, column=6).alignment = Alignment(wrap_text=True, horizontal='right', vertical='center')
                        #         ws.cell(row=new_row, column=6).hyperlink = aa[0]
                        #         ws.cell(row=new_row, column=6).value ="Image"
                        #         ws.cell(row=new_row, column=6).alignment = Alignment(wrap_text=True, horizontal='right', vertical='center')
                        #         ws.cell(row=new_row, column=6).font = dd
                        #         ws.row_dimensions[new_row].height = 150
                        #     # else:
                        #     #     ws.cell(row=new_row, column=7).value = ""


                        #     # ws.cell(row=new_row, column=7).hyperlink = nnn


                    # n =""

def accident_detail(ws,fieldnames):
    font = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    font2 = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FFFFFF')
    color = Font(name='Calibri',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='696969')
    border = Border(
            left=Side(border_style=BORDER_THIN, color='FF000000'),
            right=Side(border_style=BORDER_THIN, color='FF000000'),
            top=Side(border_style=BORDER_THIN, color='FF000000'),
            bottom=Side(border_style=BORDER_THIN, color='FF000000')
        )
    fill=PatternFill(start_color = '00C0C0C0',
            end_color = '00C0C0C0',
            fill_type = 'solid')


    row_count = ws.max_row
    column_count = ws.max_column
    for cell in ws['8:8']:
        cell.fill = fill
        cell.font = font

    for cell2 in ws['9:9']:
        cell2.fill = fill
        cell2.font = font

    for i in range(3,row_count+1):
        ws.row_dimensions[i].height = 50

    ws.row_dimensions[2].height = 40
    ws.row_dimensions[1].height = 40
    # for cell2 in ws['9:9']:
    #     cell2.fill = fill
    #     cell2.font = font

    
   
    column_widths = []
    for row in fieldnames:
        for i in range(len(row)):
            if len(column_widths) > i:
                if len(row) > column_widths[i]:
                    column_widths[i] = len(row)
            else:
                column_widths += [len(row)]
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width
    
    
    
    
    for row in ws:
        for cell1 in row:
            cell1.border = border
            cell1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    for i0 in ws['1:1']:
        i0.alignment = Alignment(horizontal='left', vertical='center')
    for i1 in ws['2:2']:
        i1.alignment = Alignment(horizontal='left', vertical='center')

    for i2 in ws['3:3']:
        i2.alignment = Alignment(horizontal='left', vertical='center')
    for i3 in ws['4:4']:
        i3.alignment = Alignment(horizontal='left', vertical='center')
    for i4 in ws['5:5']:
        i4.alignment = Alignment(horizontal='left', vertical='center')
    for i5 in ws['6:6']:
        i5.alignment = Alignment(horizontal='left', vertical='center')

def accident_report(ws, fieldnames):
    # row=30

    font = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
    font2 = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FFFFFF')
    color = Font(name='Calibri',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='696969')
    border = Border(
            left=Side(border_style=BORDER_THIN, color='FF000000'),
            right=Side(border_style=BORDER_THIN, color='FF000000'),
            top=Side(border_style=BORDER_THIN, color='FF000000'),
            bottom=Side(border_style=BORDER_THIN, color='FF000000')
        )
    fill=PatternFill(start_color = '00C0C0C0',
            end_color = '00C0C0C0',
            fill_type = 'solid')


    row_count = ws.max_row
    column_count = ws.max_column
    # for cell in ws['8:8']:
    #     cell.font = font

    # for cell2 in ws['9:9']:
    #     cell2.font = font

    # for i in range(3,row_count+1):
    #     ws.row_dimensions[i].height = 50

    # ws.row_dimensions[2].height = 40
    # ws.row_dimensions[1].height = 40
    # # for cell2 in ws['9:9']:
    #     cell2.fill = fill
    #     cell2.font = font

    
   
    column_widths = []
    for row in fieldnames:
        for i in range(len(row)):
            if len(column_widths) > i:
                if len(row) > column_widths[i]:
                    column_widths[i] = len(row)
            else:
                column_widths += [len(row)]
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width
    
    
    for row in ws:
        for cell1 in row:
            cell1.border = border
            cell1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    for i0 in ws['1:1']:
        i0.alignment = Alignment(horizontal='left', vertical='center')
    for i1 in ws['2:2']:
        i1.alignment = Alignment(horizontal='left', vertical='center')

    for i2 in ws['3:3']:
        i2.alignment = Alignment(horizontal='left', vertical='center')
    for i3 in ws['4:4']:
        i3.alignment = Alignment(horizontal='left', vertical='center')
    for i4 in ws['5:5']:
        i4.alignment = Alignment(horizontal='left', vertical='center')
    for i5 in ws['6:6']:
        i5.alignment = Alignment(horizontal='left', vertical='center')
    

    i = []
    j = []
    row = ws.max_row
    j.append(row)
    if i not in j:
        i.append(j)
    print(i)
    new_row = i[0][0] +2
    
    
    ws['A'+str(new_row)] = 'Foot Note'
    ws.merge_cells('A'+str(new_row+1)+':H'+str(new_row+1)+'')
    n2 = ws.cell(row=new_row+1,column=1)
    n2.value = "a. Urban/Rural and Details of surrounding Land use."

    ws.merge_cells('A'+str(new_row+2)+':H'+str(new_row+2)+'')
    n3 = ws.cell(row=new_row+2,column=1)
    n3.value = "b. 1. Over turning. 2. Head on collision 3. Rear End collision 4. Collision Brush/Side Wipe. 5. Left Turn collision. 6. Skidding 7. Right Turn Collision 8. Others (Pl. Specific)"

    ws.merge_cells('A'+str(new_row+3)+':H'+str(new_row+3)+'')
    n4 = ws.cell(row=new_row+3,column=1)
    n4.value = "c. 1. Fatal. 2. Grievous Injury 3. Minor Injured. 4. Non-Injury."

    ws.merge_cells('A'+str(new_row+4)+':H'+str(new_row+4)+'')
    n5 = ws.cell(row=new_row+4,column=1)
    n5.value = "d. 1. Drunken driving, 2. Over speeding, 3. Vehicle out of control, 4. Fault of driver of motor vehicle/cyclist/pedestrian/passenger, 5. Defect in mechanical /electrical condition of motor vehicle, 6. Wrong side driving, 7. Tyre bust, 8. Rain, 9. Poor visibility(fog/dust) 10. Poor Road condition 11. Others (specify)"

    ws.merge_cells('A'+str(new_row+5)+':H'+str(new_row+5)+'')
    n6 = ws.cell(row=new_row+5,column=1)
    n6.value = "e. 1. Single lane, 2. Two lanes, 3. Three lanes or more without central divider (median), 4. Four lanes or more with central divider."

    ws.merge_cells('A'+str(new_row+6)+':H'+str(new_row+6)+'')
    n7 = ws.cell(row=new_row+6,column=1)
    n7.value = "f. 1. Straight road, 2. Slight Curve, 3. Curve, 4. Flat Road, 5. Gentle incline, 6. Steep incline, 7. Hump, 8. Dip"

    ws.merge_cells('A'+str(new_row+7)+':H'+str(new_row+7)+'')
    n8 = ws.cell(row=new_row+7,column=1)
    n8.value = "g. 1. T-junction, 2. Y-junction, 3. Four arm junction, 4. Staggered junction, 5. Junction with more than 4 arms, 6. Round about junction, 7. Manned Rail crossing. 9. Unmanned Rail crossing"

    ws.merge_cells('A'+str(new_row+8)+':H'+str(new_row+8)+'')
    n9 = ws.cell(row=new_row+8,column=1)
    n9.value = "h. 1. Fine, 2. Mist/Fog, 3. Cloudy, 4-Light rain, 5. Heavy rain, 6. Hail/sleet, 7. Snow, 8. Strong Wind, 9. Dust Storm, 10. Very Hot, 11. Very Cold, 12. Other extraordinary weather condition."


def write_excel(self,request, project, queryset, type, status,start_date, end_date,asset, performance, photo,doc_type,name,accident_report_type):

    # print("-------------------------")
    # user_id = self.request.user.id
    # user = ""
    # u = User.objects.filter(id=user_id)
    # if u:
    #     user = User.objects.get(id=user_id)
    # print(user.full_name)
    
    str_start_date = datetime.strptime(start_date,'%Y-%m-%d')
    new_start_date = datetime.strftime(str_start_date,"%d-%m-%Y")

    str_end_date = datetime.strptime(end_date,'%Y-%m-%d')
    new_end_date = datetime.strftime(str_new_end_date,"%d-%m-%Y")
    


    if not request.user.is_authenticated():
        u = User.objects.filter(id=48)
        if u:
            user = User.objects.get(id=48)
           
    else:
        u = User.objects.filter(id=request.user.id)
        if u:
            user = User.objects.get(id=request.user.id)

    
    wb = Workbook()
    ws1 = wb.active
    ws2 = wb.active
    ws3 = wb.active
    ws4 = wb.active
    ws5 = wb.active
    ws6 = wb.active
    ws8 = wb.active
    queryset = queryset.prefetch_related("attachments",
                                         "generated_user_stories",
                                         "custom_attributes_values")
    queryset = queryset.select_related("owner",
                                       "assigned_to",
                                       "status",
                                       "project",
                                       "type")
    queryset = attach_total_voters_to_queryset(queryset)
    queryset = attach_watchers_to_queryset(queryset)
    if type == 'Issue' and photo=="with photo" and status==None:
        ws1.title = "Inspection Report"
        ws1['A1'] = "User Name:"
        ws1['A2'] = "User Id:"
        ws1['A3'] = "Role:"
        ws1['A4'] = "Project Name:"
        ws1['A5'] = "Report Name:"
        ws1['A6'] = "Asset Type:"
        ws1['A7'] = "Performance Parameter:"
        ws1['A8'] = "From Date: "+new_start_date
        # ws1['A1'] = "Inspection Report with Photogragh"
        # ws1['A2'] = "Project Name"

        ws1['B1'] = user.full_name
        ws1['B2'] = user.email
        # ws1['B1'] = ""
        # ws1['B2'] = ""
        ws1['B3'] = "Admin" if user.custom_role=="1" else "User"
        
        ws1.merge_cells('B4:E4')
        n1 = ws1.cell(row=4,column=2)
        n1.value = project.name
        ws1.merge_cells('B5:C5')
        n2 = ws1.cell(row=5,column=2)
        n2.value = "Inspection Report with Photograph"

        ws1['B6'] = asset if asset else "All"
        ws1['B7'] = performance if performance else "All"
        ws1['B8'] = "To Date: "+new_end_date
        # ws1['B2'] = project.name

        fieldnames = ["Ref.No.", "Chainage","" , "Direction", "Description of Issue",
                              "Photograph During Inspection", "Asset Type", "Performance Parameter",
                              "Issue Raised On", "Issue Raised By",
                              "Issue Assigned To"]
        ws1.append(fieldnames)
        ws1.merge_cells('A9:A10')
        ws1.merge_cells('B9:C9')
        # ws1.merge_cells('C3:D3')
        # ws1.merge_cells('C4:D4')
        n1 = ws1.cell(row=10,column=2)
        n2 = ws1.cell(row=10,column=3)
        n1.value = "From (In Km)"
        n2.value = "To (In Km)"
        ws1.merge_cells('D9:D10')
        # n3 = ws1.cell(row=3,column=5)
        # n3.value="Direction"
        ws1.merge_cells('E9:E10')
        ws1.merge_cells('F9:F10')
        ws1.merge_cells('G9:G10')
        ws1.merge_cells('H9:H10')
        ws1.merge_cells('I9:I10')
        ws1.merge_cells('J9:J10')
        ws1.merge_cells('K9:K10')
        ws1.merge_cells('L9:L10')
        

    if type == 'Issue' and photo=="without photo" and status==None:
        ws5.title = "Inspection Reportssssss"
        ws5['A1'] = "User Name:"
        ws5['A2'] = "User Id:"
        ws5['A3'] = "Role:"
        ws5['A4'] = "Project Name:"
        ws5['A5'] = "Report Name:"
        ws5['A6'] = "Asset Type:"
        ws5['A7'] = "Performance Parameter:"
        ws5['A8'] = "From Date: "+new_start_date
        # ws1['A1'] = "Inspection Report with Photogragh"
        # ws1['A2'] = "Project Name"

        ws1['B1'] = user.full_name
        ws1['B2'] = user.email
        # ws1['B1'] = ""
        # ws1['B2'] = ""
        ws5['B3'] = "Admin" if user.custom_role=="1" else "User"
        
        ws5.merge_cells('B4:E4')
        n1 = ws5.cell(row=4,column=2)
        n1.value = project.name
        
        ws5.merge_cells('B5:C5')
        n2 = ws5.cell(row=5,column=2)
        n2.value = "Inspection Report without Photograph"

        ws5['B6'] = asset if asset else "All"
        ws5['B7'] = performance if performance else "All"
        ws5['B8'] ="To Date: "+ new_end_date

        fieldnames = ["Ref.No.", "Chainage","" , "Direction", "Description of Issue",
                              "Asset Type", "Performance Parameter",
                              "Issue Raised On", "Issue Raised By",
                              "Issue Assigned To"]
        ws5.append(fieldnames)
    
        ws5.merge_cells('A9:A10')
        ws5.merge_cells('B9:C9')
        # ws1.merge_cells('C3:D3')
        # ws1.merge_cells('C4:D4')
        n1 = ws5.cell(row=10,column=2)
        n2 = ws5.cell(row=10,column=3)
        n1.value = "From (In Km)"
        n2.value = "To (In Km)"
        ws5.merge_cells('D9:D10')
        # n3 = ws1.cell(row=3,column=5)
        # n3.value="Direction"
        ws5.merge_cells('E9:E10')
        ws5.merge_cells('F9:F10')
        ws5.merge_cells('G9:G10')
        ws5.merge_cells('H9:H10')
        ws5.merge_cells('I9:I10')
        ws5.merge_cells('J9:J10')
        # ws1.merge_cells('L3:L4')
    if type == 'Issue' and name=="Compliance" and photo=="with photo" and status:

        wb = Workbook()
        ws2 = wb.active
    
        ws2.title = "Manitenance Report"
        
        ws2['A1'] = "User Name:"
        ws2['A2'] = "User Id:"
        ws2['A3'] = "Role:"
        ws2['A4'] = "Project Name:"
        ws2['A5'] = "Report Name:"
        ws2['A6'] = "Asset Type:"
        ws2['A7'] = "Performance Parameter:"
        ws2['A8'] = "From Date: "+new_start_date
        # ws1['A1'] = "Inspection Report with Photogragh"
        # ws1['A2'] = "Project Name"

        ws2['B1'] = user.full_name
        ws2['B2'] = user.email
        # ws1['B1'] = ""
        # ws1['B2'] = ""
        ws2['B3'] = "Admin" if user.custom_role=="1" else "User"
        
        ws2.merge_cells('B4:E4')
        n1 = ws2.cell(row=4,column=2)
        n1.value = project.name
        
        ws2.merge_cells('B5:C5')
        n2 = ws2.cell(row=5,column=2)
        n2.value = "Manitenance Report with Photograph"

        ws2['B6'] = asset if asset else "All"
        ws2['B7'] = performance if performance else "All"
        ws2['B8'] = "To Date: "+new_end_date

        fieldnames = ["Ref.No.", "Chainage","", "Direction", "Description of Issue",
                          "Photograph During Inspection","Photograph During Maintenance", "Asset Type", "Performance Parameter",
                          "Issue Raised On", "Issue Raised By",
                          "Issue Assigned To" , "Max Time limit for Rectification/Repair",
                          "", "Action Taken",
                          "", "", "Issue Closed By","Description Of Compliance",
                          "Remark", "Current Status"]
        ws2.append(fieldnames)
        ws2.merge_cells('A9:A10')
        ws2.merge_cells('B9:C9')
        # ws2.merge_cells('C3:D3')
        # ws1.merge_cells('C4:D4')
        n1 = ws2.cell(row=10,column=2)
        n2 = ws2.cell(row=10,column=3)
        n1.value = "From (In Km)"
        n2.value = "To (In Km)"
        ws2.merge_cells('D9:D10')
        # n3 = ws1.cell(row=3,column=5)
        # n3.value="Direction"
        ws2.merge_cells('E9:E10')
        ws2.merge_cells('F9:F10')
        ws2.merge_cells('G9:G10')
        ws2.merge_cells('H9:H10')
        ws2.merge_cells('I9:I10')
        ws2.merge_cells('J9:J10')
        ws2.merge_cells('K9:K10')
        ws2.merge_cells('L9:L10')
        ws2.merge_cells('M9:N9')
        n1 = ws2.cell(row=10,column=13)
        n2 = ws2.cell(row=10,column=14)
        n1.value = "Timeline\n (As per  Schedule F)"
        n2.value = "Target Date\n(As per  Schedule F)"
        # ws2.merge_cells('M3:M4')
        # ws2.merge_cells('N3:N4')
        ws2.merge_cells('O9:Q9')
        n1 = ws2.cell(row=10,column=15)
        n2 = ws2.cell(row=10,column=16)
        n3 = ws2.cell(row=10,column=17)
        n1.value = "Status"
        n2.value = "Issue Closed On Date"
        n3.value = "Complianced"

        # ws2.merge_cells('Q3:Q4')
        ws2.merge_cells('R9:R10')
        ws2.merge_cells('S9:S10')
        ws2.merge_cells('T9:T10')
        ws2.merge_cells('U9:U10')


    if type=='Issue' and name=="Compliance" and photo=="without photo" and status:
        ws4.title = "Manitenance Report"
        ws4['A1'] = "User Name:"
        ws4['A2'] = "User Id:"
        ws4['A3'] = "Role:"
        ws4['A4'] = "Project Name:"
        ws4['A5'] = "Report Name:"
        ws4['A6'] = "Asset Type:"
        ws4['A7'] = "Performance Parameter:"
        ws4['A8'] = "From Date: "+new_start_date
        # ws1['A1'] = "Inspection Report with Photogragh"
        # ws1['A2'] = "Project Name"

        ws4['B1'] = user.email
        ws4['B2'] = user.full_name
        # ws1['B1'] = ""
        # ws1['B2'] = ""
        ws4['B3'] = "Admin" if user.custom_role=="1" else "User"
        
        ws4.merge_cells('B4:E4')
        n1 = ws4.cell(row=4,column=2)
        n1.value = project.name
        
        ws4.merge_cells('B5:C5')
        n2 = ws4.cell(row=5,column=2)
        n2.value = "Manitenance Report without Photograph"

       
        ws4['B6'] = asset if asset else "All"
        ws4['B7'] = performance if performance else "All"
        ws4['B8'] = "To Date: "+new_end_date

        fieldnames = ["Ref.No.", "Chainage","", "Direction", "Description of Issue",
                          "Asset Type", "Performance Parameter",
                          "Issue Raised On", "Issue Raised By",
                          "Issue Assigned To" , "Max Time limit for Rectification/Repair",
                          "", "Action Taken",
                          "", "", "Issue Closed By","Description Of Compliance",
                           "Remark", "Current Status"]
        ws4.append(fieldnames)
        ws4.merge_cells('A9:A10')
        ws4.merge_cells('B9:C9')
        # ws2.merge_cells('C3:D3')
        # ws1.merge_cells('C4:D4')
        n1 = ws4.cell(row=10,column=2)
        n2 = ws4.cell(row=10,column=3)
        n1.value = "From (In Km)"
        n2.value = "To (In Km)"
        ws4.merge_cells('D9:D10')
        # n3 = ws1.cell(row=3,column=5)
        # n3.value="Direction"
        ws4.merge_cells('E9:E10')
        ws4.merge_cells('F9:F10')
        ws4.merge_cells('G9:G10')
        ws4.merge_cells('H9:H10')
        ws4.merge_cells('I9:I10')
        ws4.merge_cells('J9:J10')
        # ws4.merge_cells('K3:K4')
        # ws2.merge_cells('L3:L4')
        ws4.merge_cells('K9:L9')
        n1 = ws4.cell(row=10,column=11)
        n2 = ws4.cell(row=10,column=12)
        n1.value = "Timeline"
        n2.value = "Target Date"
        # ws2.merge_cells('M3:M4')
        # ws2.merge_cells('N3:N4')
        ws4.merge_cells('M9:O9')
        n1 = ws4.cell(row=10,column=13)
        n2 = ws4.cell(row=10,column=14)
        n3 = ws4.cell(row=10,column=15)
        n1.value = "Status"
        n2.value = "Issue Closed On Date"
        n3.value = "Complianced"
        ws4.merge_cells('P9:P10')
        ws4.merge_cells('Q9:Q10')
        ws4.merge_cells('R9:R10')
        ws4.merge_cells('S9:S10')


    if type == 'Investigation' and photo=="with photo":
        
        ws3.title = "Test Report"
        ws3['A1'] = "User Name:"
        ws3['A2'] = "User Id:"
        ws3['A3'] = "Role:"
        ws3['A4'] = "Project Name:"
        ws3['A5'] = "Report Name:"
        ws3['A6'] = "Asset Type:"
        ws3['A7'] = "Performance Parameter:"
        ws3['A8'] = "From Date: "+new_start_date
        # ws1['A1'] = "Inspection Report with Photogragh"
        # ws1['A2'] = "Project Name"

        ws3['B1'] = user.full_name
        ws3['B2'] = user.email
        # ws1['B1'] = ""
        # ws1['B2'] = ""
        ws3['B3'] = "Admin" if user.custom_role=="1" else "User"
    
        ws3.merge_cells('B4:E4')
        n1 = ws3.cell(row=4,column=2)
        n1.value = project.name
        
        ws3.merge_cells('B5:C5')
        n2 = ws3.cell(row=5,column=2)
        n2.value = "Test and Investigation Report"

       
        ws3['B6'] = asset if asset else "All"
        ws3['B7'] = performance if performance else "All"
        ws3['B8'] = "To Date: "+new_end_date

        fieldnames = ["Ref.No.","Description of Test/ Investigation", "Chainage","", "Direction",
                          "Asset Type", "Performance Parameter",
                          "Name of Test", "Testing Method", "Standard References for testing",
                          "Test Carried Out Date", "Testing Carried Out By(Name)"]
        
        ws3.append(fieldnames)
        ws3.merge_cells('A9:A10')
        ws3.merge_cells('B9:B10')
        ws3.merge_cells('C9:D9')
        # ws1.merge_cells('C4:D4')
        n1 = ws3.cell(row=10,column=3)
        n2 = ws3.cell(row=10,column=4)
        n1.value = "From (In Km)"
        n2.value = "To (In Km)"
        # ws1.merge_cells('D3:D4')
        # n3 = ws1.cell(row=3,column=5)
        # n3.value="Direction"
        ws3.merge_cells('E9:E10')
        ws3.merge_cells('F9:F10')
        ws3.merge_cells('G9:G10')
        ws3.merge_cells('H9:H10')
        ws3.merge_cells('I9:I10')
        ws3.merge_cells('J9:J10')
        ws3.merge_cells('K9:K10')
        ws3.merge_cells('L9:L10')
  
    if type == 'Investigation' and photo=="without photo":
        
        ws6.title = "Test Report"
        ws6['A1'] = "User Name:"
        ws6['A2'] = "User Id:"
        ws6['A3'] = "Role:"
        ws6['A4'] = "Project Name:"
        ws6['A5'] = "Report Name:"
        ws6['A6'] = "Asset Type:"
        ws6['A7'] = "Performance Parameter:"
        ws6['A8'] = "From Date: "+new_start_date
        # ws1['A1'] = "Inspection Report with Photogragh"
        # ws1['A2'] = "Project Name"

        ws6['B1'] = user.full_name
        ws6['B2'] = user.email
        # ws1['B1'] = ""
        # ws1['B2'] = ""
        ws6['B3'] = "Admin" if user.custom_role=="1" else "User"

        ws6.merge_cells('B4:E4')
        n1 = ws6.cell(row=4,column=2)
        n1.value = project.name
        
        ws6.merge_cells('B5:C5')
        n2 = ws6.cell(row=5,column=2)
        n2.value = "Test and Investigation Report"
       
        ws6['B6'] = asset if asset else "All"
        ws6['B7'] = performance if performance else "All"
        ws6['B8'] = "To Date: "+new_end_date

        fieldnames = ["Ref.No.","Description of Test/ Investigation", "Chainage","", "Direction",
                          "Asset Type", "Performance Parameter",
                          "Name of Test", "Testing Method", "Standard References for testing",
                          "Test Carried Out Date", "Testing Carried Out By(Name)"]
        
        ws6.append(fieldnames)
        ws6.merge_cells('A9:A10')
        ws6.merge_cells('B9:B10')
        ws6.merge_cells('C9:D9')
        # ws1.merge_cells('C4:D4')
        n1 = ws6.cell(row=10,column=3)
        n2 = ws6.cell(row=10,column=4)
        n1.value = "From (In Km)"
        n2.value = "To (In Km)"
        # ws1.merge_cells('D3:D4')
        # n3 = ws1.cell(row=3,column=5)
        # n3.value="Direction"
        ws6.merge_cells('E9:E10')
        ws6.merge_cells('F9:F10')
        ws6.merge_cells('G9:G10')
        ws6.merge_cells('H9:H10')
        ws6.merge_cells('I9:I10')
        ws6.merge_cells('J9:J10')
        ws6.merge_cells('K9:K10')
        ws6.merge_cells('L9:L10')
  



    if type == 'Accident' and accident_report_type== "Summary":
        ws4.title = "Summary of Accident"
        ws4['A1'] = "User Name:"
        ws4['A2'] = "User Id:"
        ws4['A3'] = "Role:"
        ws4['A4'] = "Project Name:"
        ws4['A5'] = "Report Name:"
        ws4['A6'] = "From Date: "+new_start_date
        ws4['A7'] = ""
        # ws1['A1'] = "Inspection Report with Photogragh"
        # ws1['A2'] = "Project Name"

        ws4['B1'] = user.full_name
        ws4['B2'] = user.email
        # ws1['B1'] = ""
        # ws1['B2'] = ""
        ws4['B3'] = "Admin" if user.custom_role=="1" else "User"
        
        ws4.merge_cells('B4:E4')
        n1 = ws4.cell(row=4,column=2)
        n1.value = project.name
        
        ws4.merge_cells('B5:C5')
        n2 = ws4.cell(row=5,column=2)
        n2.value = "Summary of Accident Report"
       
        ws4['B6'] = "To Date: "+new_end_date
        ws4['B7'] = ""
        fieldnames = ["Ref.No.", "Description","Up to previous month","","During this month","",
                        "Up to this month", ""]

        ws4.append(fieldnames)
        ws4.merge_cells('A8:A9')
        ws4.merge_cells('B8:B9')
        ws4.merge_cells('C8:D8')
        n1 = ws4.cell(row=9,column=3)
        n2 = ws4.cell(row=9,column=4)
        n1.value = "No of Accidents"
        n2.value = "No of Peoples affected"
        ws4.merge_cells('E8:F8')
        n1 = ws4.cell(row=9,column=5)
        n2 = ws4.cell(row=9,column=6)
        n1.value = "No of Accidents"
        n2.value = "No of Peoples affected"
        ws4.merge_cells('G8:H8')
        n1 = ws4.cell(row=9,column=7)
        n2 = ws4.cell(row=9,column=8)
        n1.value = "No of Accidents"
        n2.value = "No of Peoples affected"
    
    if type == 'Accident' and accident_report_type== "Detail":
        ws8.title = "Detail of Accident"
        ws8['A1'] = "User Name:"
        ws8['A2'] = "User Id:"
        ws8['A3'] = "Role:"
        ws8['A4'] = "Project Name:"
        ws8['A5'] = "Report Name:"
        ws8['A6'] = "From Date: "+new_start_date
        ws8['A7'] = ""
        # ws1['A1'] = "Inspection Report with Photogragh"
        # ws1['A2'] = "Project Name"

        ws8['B1'] = user.full_name
        ws8['B2'] = user.email
        # ws1['B1'] = ""
        # ws1['B2'] = ""
        ws8['B3'] = "Admin" if user.custom_role=="1" else "User"    

        ws8.merge_cells('B4:E4')
        n1 = ws8.cell(row=4,column=2)
        n1.value = project.name
        
        ws8.merge_cells('B5:C5')
        n2 = ws8.cell(row=5,column=2)
        n2.value = "Detail of Accident Report"
       
        ws8['B6'] = "To Date: "+new_end_date
        ws8['B7'] = ""
        # ws8['A'] = "Foot Note"

        

        fieldnames = ["Sr. No","Date","Time of Accident","Accident Location","Nature of Accident",
                        "Classification of Accident","Causes","Road Feature","Road Condition",
                        "Intersection type and control","Weather Condition","Vehicle Responsible",
                        "Fatal","Grievous","Minor","Non Injured","No. of Animals killed if any",
                        "Help provided by Ambulance/ RPV/ Crane"]

        

        ws8.merge_cells('A8:R8')
        n2 = ws8.cell(row=8,column=1)
        n2.value = "ACCIDENT DATA FROM "+new_start_date+" TO "+new_end_date
        ws8.merge_cells('A9:C9')
        n1 = ws8.cell(row=9,column=1)
        n1.value = "Legends as per Foot Note"
        n1 = ws8.cell(row=9,column=4)
        n1.value = "a"
        n1 = ws8.cell(row=9,column=5)
        n1.value = "b"
        n1 = ws8.cell(row=9,column=6)
        n1.value = "c"
        n1 = ws8.cell(row=9,column=7)
        n1.value = "d"
        n1 = ws8.cell(row=9,column=8)
        n1.value = "e"
        n1 = ws8.cell(row=9,column=9)
        n1.value = "f"
        n1 = ws8.cell(row=9,column=10)
        n1.value = "g"
        n1 = ws8.cell(row=9,column=11)
        n1.value = "h"
        # ws8.merge_cells('A28:H28')
        # n9 = ws8.cell(row=28, column=1)
        
        ws8.merge_cells('M9:P9')
        n1 = ws8.cell(row=9,column=13)
        n1.value = "No. of affected person"
        ws8.append(fieldnames)
        

    count = 0
    for issue in queryset:
        count += 1 
        if issue.type.name=='Issue' and photo=="with photo" and status==None:
            qqq = issue.watchers
            watchers = []
            new_watcher_list =  ""
            watcher_username =""
            if issue.assigned_to:
                watcher_username = '1. '+issue.assigned_to.full_name 
            for i in qqq:
                sql = User.objects.get(id=int(i))
                watchers.append(sql.full_name)
            for j in range(len(watchers)):
                watcher_username = str(j+2)+'. '+watchers[j] +','+ watcher_username
            
            split = watcher_username.split(',')

            for i in range(len(split)):
                new_watcher_list = split[i]+'\n' + new_watcher_list
            if issue.attachments:
                file_name = "" 
                files = []
                file = issue.attachments.all().filter(project__id=issue.project.id,description="").values_list('attached_file')
                for i in file:
                    files.extend(i)
                for j in files:
                    file_name = os.path.join(settings.MEDIA_URL,str(j)) +'\n' + file_name
            else:
                file_name=""

            print(file_name)
            
            Raised_date = datetime.strftime(issue.created_date.date(),"%d-%m-%Y")

            issue_data = [[
                issue.ref,
                # issue.project.name,
                issue.chainage_from,
                issue.chainage_to,
                issue.chainage_side,
                issue.description,
                file_name if issue.attachments else None,
                issue.issue_category,
                issue.issue_subcategory,
                Raised_date,
                issue.owner.full_name if issue.owner else None,
                new_watcher_list,
            ]]
            for data in issue_data:
                ws1.append(data)
            wb.save("table.xlsx")
            wb.close()

            wb = load_workbook('table.xlsx')
            ws1 = wb['Inspection Report']

            style(ws1,fieldnames, file_name, issue)
            wb.save("table.xlsx")


        if issue.type.name == 'Issue' and photo=="without photo" and status==None:
            qqq = issue.watchers
            watchers = []
            new_watcher_list =  ""
            watcher_username =""
            if issue.assigned_to:
                watcher_username = '1. '+issue.assigned_to.full_name 
            for i in qqq:
                sql = User.objects.get(id=int(i))
                watchers.append(sql.full_name)
            for j in range(len(watchers)):
                watcher_username = str(j+2)+'. '+watchers[j] +','+ watcher_username
            
            split = watcher_username.split(',')

            for i in range(len(split)):
                new_watcher_list = split[i]+'\n' + new_watcher_list
                
            Raised_date = datetime.strftime(issue.created_date.date(),"%d-%m-%Y")
            issue_data = [[
                issue.ref,
                # issue.project.name,
                issue.chainage_from,
                issue.chainage_to,
                issue.chainage_side,
                issue.description,
                issue.issue_category,
                issue.issue_subcategory,
                Raised_date,
                issue.owner.full_name if issue.owner else None,
                new_watcher_list,
            ]]

            for data in issue_data:
                ws5.append(data)
            wb.save("table.xlsx")
            wb.close()

            wb = load_workbook('table.xlsx')
            ws5 = wb['Inspection Reportssssss']
            style(ws5,fieldnames, issue)
            wb.save("table.xlsx")
    

        if issue.type.name=='Issue' and name=="Compliance" and photo=="with photo" and status:
            
            qqq = issue.watchers
            watchers = []
            new_watcher_list =  ""
            watcher_username =""
            if issue.assigned_to:
                watcher_username = '1. '+issue.assigned_to.full_name 
            for i in qqq:
                sql = User.objects.get(id=int(i))
                watchers.append(sql.full_name)
            for j in range(len(watchers)):
                watcher_username = str(j+2)+'. '+watchers[j] +','+ watcher_username
            
            split = watcher_username.split(',')

            for i in range(len(split)):
                new_watcher_list = split[i]+'\n' + new_watcher_list

            a = issue.created_date.date()
            b = datetime.strptime(issue.target_date,"%d/%m/%Y").date()
            timeline = b-a
            days = timeline.days
            timeline_hrs = str(days*24) +" Hrs"
            
            target_date = datetime.strftime(b,"%d-%m-%Y")

            c = issue.modified_date.date()
            
            modified_date = datetime.strftime(c,"%d-%m-%Y")

            if issue.attachments:
                file_name = "" 
                files = []
                Compliance_file_name = "" 
                Compliance_files = []

                file = issue.attachments.all().filter(project__id=issue.project.id,description="").values_list('attached_file')
                if file:
                    for i in file:
                        files.extend(i)
                for j in files:
                    file_name = os.path.join(settings.MEDIA_URL,str(j)) +'\n' + file_name
            

                
                Compliance_file = issue.attachments.all().filter(project__id=issue.project.id,description="Compliances").values_list('attached_file')
                if Compliance_file:
                    for k in Compliance_file:
                        Compliance_files.extend(k)
                for l in Compliance_files:
                    Compliance_file_name = os.path.join(settings.MEDIA_URL,str(l)) +'\n' + Compliance_file_name
            else:
                file_name=""
                Compliance_file_name=""
          
            status_name = []
            status_names =  project.issues.filter(status__id__in=status)
            new_status_name =[]
            for i in status_names:
                
                if str(i.status) == 'Closed':
                    new_status_name.append('Open')
                    # new_status_name += 'Open'
                     
                elif str(i.status) == 'Maintenance Closed':
                    # new_status_name += 'Closed'
                    new_status_name.append('Closed')
                    
                elif str(i.status) == 'Maintenance Pending':
                    # new_status_name += 'Pending'
                    new_status_name.append('Pending')
            if new_status_name:
                new =""
                for i in new_status_name:
                    new = i

            Raised_date = datetime.strftime(issue.created_date.date(),"%d-%m-%Y")
            issue_data = [[
                    issue.ref,
                    issue.chainage_from,
                    issue.chainage_to,
                    issue.chainage_side,
                    issue.description,
                    file_name,
                    Compliance_file_name,
                    issue.issue_category,
                    issue.issue_subcategory,
                    Raised_date,
                    issue.owner.full_name if issue.owner else None,
                    new_watcher_list,
                    timeline_hrs,
                    target_date,
                    str(issue.status) if issue.status else None,
                    modified_date if str(issue.status)=='Closed' else None,
                    'Yes' if issue.compliance_is_update==True else 'No',
                    issue.assigned_to.full_name if issue.assigned_to else None,
                    issue.compliance_description,
                    "",
                    str(issue.status) if issue.status else None,
                ]]
            for data in issue_data:
                ws2.append(data)
            wb.save("table.xlsx")
            wb.close()

            wb = load_workbook('table.xlsx')
            ws2 = wb['Manitenance Report']
            style(ws2,fieldnames, file_name, issue)

            comp(ws2,Compliance_file_name)
            wb.save("table.xlsx")

        if issue.type.name=='Issue' and name=="Compliance" and photo=="without photo" and status:
            qqq = issue.watchers
            watchers = []
            new_watcher_list =  ""
            watcher_username =""
            if issue.assigned_to:
                watcher_username = '1. '+issue.assigned_to.full_name 
            for i in qqq:
                sql = User.objects.get(id=int(i))
                watchers.append(sql.full_name)
            for j in range(len(watchers)):
                watcher_username = str(j+2)+'. '+watchers[j] +','+ watcher_username
            
            split = watcher_username.split(',')

            for i in range(len(split)):
                new_watcher_list = split[i]+'\n' + new_watcher_list
            a = issue.created_date.date()
            b = datetime.strptime(issue.target_date,"%d/%m/%Y").date()
            timeline = b-a
            target_date = datetime.strftime(b,"%d-%m-%Y")
            
            status_name = []
            status_names =  project.issues.filter(status__id__in=status)
            new_status_name =[]
            for n in status_names:
                
                if str(n.status) == 'Closed':
                    new_status_name.append('Open')
                    # new_status_name += 'Open'
                     
                elif str(n.status) == 'Maintenance Closed':
                    # new_status_name += 'Closed'
                    new_status_name.append('Closed')
                    
                elif str(n.status) == 'Maintenance Pending':
                    # new_status_name += 'Pending'
                    new_status_name.append('Pending')
            if new_status_name:
                new =""
                for i in new_status_name:
                    new = i
            Raised_date = datetime.strftime(issue.created_date.date(),"%d-%m-%Y")
            issue_data = [[
                    issue.ref,
                    issue.chainage_from,
                    issue.chainage_to,
                    issue.chainage_side,
                    issue.description,
                    issue.issue_category,
                    issue.issue_subcategory,
                    Raised_date,
                    issue.owner.full_name if issue.owner else None,
                    new_watcher_list,
                    timeline,
                    target_date,
                    str(issue.status) if issue.status else None,
                    issue.finished_date if status_name=='Closed' else None,
                    'Yes' if issue.compliance_is_update==True else 'No',
                    issue.assigned_to.full_name if issue.assigned_to else None,
                    issue.compliance_description,
                    "",
                    str(issue.status) if issue.status else None,
                ]]
            for data in issue_data:
                ws4.append(data)
            wb.save("table.xlsx")
            wb.close()

            wb = load_workbook('table.xlsx')
            ws4 = wb['Manitenance Report']
            style(ws4,fieldnames, issue)
            wb.save("table.xlsx")

        if issue.type.name == 'Investigation' and photo=="with photo":
            Raised_date = datetime.strftime(issue.created_date.date(),"%d-%m-%Y")
            issue_data = [[
                issue.ref,
                issue.investigation_description,
                issue.investigation_chainage_from,
                issue.investigation_chainage_to,
                issue.investigation_chainage_side,
                issue.asset_name,
                issue.test_name,
                issue.test_name,
                issue.testing_method,
                issue.test_specifications,
                Raised_date,
                issue.owner.full_name if issue.owner else None,
            ]]
            for data in issue_data:
                ws3.append(data)


            wb.save("table.xlsx")
            wb.close()

            wb = load_workbook('table.xlsx')
        
            ws3 = wb['Test Report']
            style(ws3,fieldnames, issue)

        if issue.type.name == 'Investigation' and photo=="without photo":
            Raised_date = datetime.strftime(issue.created_date.date(),"%d-%m-%Y")
            issue_data = [[
                issue.ref,
                issue.investigation_description,
                issue.investigation_chainage_from,
                issue.investigation_chainage_to,
                issue.investigation_chainage_side,
                issue.asset_name,
                issue.test_name,
                issue.test_name,
                issue.testing_method,
                issue.test_specifications,
                Raised_date,
                issue.owner.full_name if issue.owner else None,
            ]]
            for data in issue_data:
                ws6.append(data)


            wb.save("table.xlsx")
            wb.close()

            wb = load_workbook('table.xlsx')
        
            ws6 = wb['Test Report']
            style(ws6,fieldnames, issue)
            wb.save("table.xlsx")

        if issue.type.name == 'Accident' and accident_report_type== "Summary":
            
            last_day_of_prev_month = date.today().replace(day=1) - timedelta(days=1)
            first_date_of_previos_month = date.today().replace(day=1) - timedelta(days=last_day_of_prev_month.day)
            first_date = date.today().replace(day=1)
            current_date = date.today()
            previous_month = first_date_of_previos_month
            Previous_last_date = last_day_of_prev_month
            animals_killed_last_month = project.issues.filter(created_date__date__range=[previous_month,Previous_last_date],type__name='Accident').values_list('animals_killed', flat=True)
            
            animal_list_last_month = list(animals_killed_last_month)
            new_list_last = []
            
            if animals_killed_last_month:
                for i in animals_killed_last_month:
                    if i:
                        new_list_last.append(int(i))


            animals_killed_cuurent_month = project.issues.filter(created_date__date__range=[first_date,current_date],type__name='Accident').values_list('animals_killed', flat=True)
            animal_list_current_month = list(animals_killed_cuurent_month)
            new_list_current = []
            if animals_killed_cuurent_month:
                for i in animal_list_current_month:
                    if i:
                        new_list_current.append(int(i))


            animals_killed_upto_month = project.issues.filter(type__name='Accident').values_list('animals_killed', flat=True)

            animal_list_upto_month = list(animals_killed_upto_month)
            new_list_upto = []
            if animals_killed_upto_month:
                for i in animal_list_upto_month:
                    if i:
                        new_list_upto.append(int(i))
            issue_data = [[
                issue.ref,
                issue.accident_classification,
                project.issues.filter(type__name='Accident',created_date__date__range=[previous_month,Previous_last_date]).count(),
                sum(new_list_last),
                project.issues.filter(type__name='Accident',created_date__date__range=[first_date,current_date]).count(),
                sum(new_list_current),
                project.issues.filter(type__name='Accident').count(),
                sum(new_list_upto),

            ]]
            for data in issue_data:
                ws4.append(data)


            wb.save("table.xlsx")
            wb.close()

            wb = load_workbook('table.xlsx')
        
            ws4 = wb['Summary of Accident']
            # style(ws4,fieldnames, issue)
            accident_detail(ws4, fieldnames)
            wb.save("table.xlsx")

        if issue.type.name == 'Accident' and accident_report_type== "Detail":
            animals_killed_cuurent_month = project.issues.filter(created_date__date__range=[start_date,end_date],type__name='Accident').values_list('animals_killed', flat=True)
            animal_list_current_month = list(animals_killed_cuurent_month)
            new_list_current = []
            if animals_killed_cuurent_month:
                for i in animal_list_current_month:
                    if i:
                        new_list_current.append(int(i))

            
            if issue.accident_nature:
                b = 0
                if issue.accident_nature=="Over turning":
                    b = 1
                elif issue.accident_nature=="Head on collision":
                    b = 2
                elif issue.accident_nature=="Rear End collision":
                    b = 3
                elif issue.accident_nature=="Collision Brush/Side Wipe":
                    b = 4
                elif issue.accident_nature=="Left Turn collision":
                    b = 5
                elif issue.accident_nature=="Skidding":
                    b = 6
                elif issue.accident_nature=="Right Turn Collision":
                    b = 7
                elif issue.accident_nature=="Others":
                    b = 8
                else:
                    pass
            else:
                pass

            if issue.accident_classification:
                c = 0
                if issue.accident_classification=="Fatal":
                    c = 1
                elif issue.accident_classification=="Grievous Injury":
                    c = 2
                elif issue.accident_classification=="Minor Injured":
                    c = 3
                elif issue.accident_classification=="Non-Injury":
                    c = 4
                else:
                    pass
            else:
                pass

            if issue.accident_causes:
                d = 0 
                if issue.accident_causes == "Drunken driving":
                    d = 1
                elif issue.accident_causes == "Over speeding":
                    d = 2
                elif issue.accident_causes == "Vehicle out of control":
                    d = 3
                elif issue.accident_causes == "Fault of driver of motor vehicle/cyclist/pedestrian/passenger":
                    d = 4
                elif issue.accident_causes == "Defect in mechanical /electrical condition of motor vehicle":
                    d = 5
                elif issue.accident_causes == "Wrong side driving":
                    d = 6 
                elif issue.accident_causes == "Tyre bust":
                    d = 7 
                elif issue.accident_causes == "Rain":
                    d = 8
                elif issue.accident_causes == "Poor visibility(fog/dust)":
                    d = 9
                elif issue.accident_causes == "Poor Road condition":
                    d = 10
                elif issue.accident_causes == "Others":
                    d = 11
                else:
                    pass
            else:
                pass

            if issue.road_feature:
                e = 0
                if issue.road_feature=="Single lane":
                    e = 1
                elif issue.road_feature=="Two lanes":
                    e = 2
                elif issue.road_feature=="Three lanes or more without central divider (median)":
                    e = 3
                elif issue.road_feature=="Four lanes or more with central divider":
                    e = 4
                else:
                    pass
            else:
                pass

            if issue.road_condition:
                f = 0
                if issue.road_condition=="Straight road":
                    f = 1
                elif issue.road_condition=="Slight Curve":
                    f = 2
                elif issue.road_condition=="Curve":
                    f = 3
                elif issue.road_condition=="Flat Road":
                    f = 4
                elif issue.road_condition=="Gentle incline":
                    f = 5
                elif issue.road_condition=="Steep incline":
                    f = 6
                elif issue.road_condition=="Hump":
                    f = 7
                elif issue.road_condition=="Dip":
                    f = 8
                else:
                    pass
            else:
                pass
            
            if issue.intersection_type:
                g = 0
                if issue.intersection_type=="T-junction":
                    g = 1
                elif issue.intersection_type=="Y-junction":
                    g = 2
                elif issue.intersection_type=="Four arm junction":
                    g = 3
                elif issue.intersection_type=="Staggered junction":
                    g = 4
                elif issue.intersection_type=="Junction with more than":
                    g = 5
                elif issue.intersection_type=="arms":
                    g = 6
                elif issue.intersection_type=="Round about junction":
                    g = 7
                elif issue.intersection_type=="Manned Rail crossing":
                    g = 8
                elif issue.intersection_type=="Unmanned Rail crossing":
                    g = 9
                else:
                    pass
            else:
                pass
            
            if issue.weather_condition:
                h = 0
                if issue.weather_condition=="Fine":
                    h = 1
                elif issue.weather_condition=="Mist/Fog":
                    h = 2
                elif issue.weather_condition=="Cloudy":
                    h = 3
                elif issue.weather_condition=="Light rain":
                    h = 4
                elif issue.weather_condition=="Heavy rain":
                    h = 5
                elif issue.weather_condition=="Hail/sleet":
                    h = 6
                elif issue.weather_condition=="Snow":
                    h = 7
                elif issue.weather_condition=="Strong Wind":
                    h = 8
                elif issue.weather_condition=="Dust Storm":
                    h = 9
                elif issue.weather_condition=="Very Hot":
                    h = 10
                elif issue.weather_condition=="Very Cold":
                    h = 11
                elif issue.weather_condition=="Other extraordinary weather condition":
                    h = 12
                else:
                    pass
            else:
                pass




            issue_data = [[
               count,
               issue.accident_date,
               issue.accident_time,
               issue.chainage_from,
               b if issue.accident_nature else None,
               c if issue.accident_classification else None,
               d if issue.accident_causes else None, 
               e if issue.road_feature else None,
               f if issue.road_condition else None,
               g if issue.intersection_type else None,
               h if issue.weather_condition else None,
               issue.vehicle_responsible,
               issue.affected_persons_fatal,
               issue.affected_persons_grievous,
               issue.affected_persons_minor,
               issue.affected_persons_non_injured,
               sum(new_list_current),
               issue.help_provided

            ]]

            for data in issue_data:
                ws8.append(data)
            
            wb.save("table.xlsx")
            wb.close()

            wb = load_workbook('table.xlsx')
        
            ws8 = wb['Detail of Accident']

            # accident_detail(ws8, fieldnames)
            wb.save("table.xlsx")

    if type=="Accident" and accident_report_type== "Detail":
        wb = load_workbook('table.xlsx')
        
        ws8 = wb['Detail of Accident']

        accident_report(ws8, fieldnames)
        wb.save("table.xlsx")
          
    if doc_type=="pdf":
        new = pd.read_excel('table.xlsx',na_filter=False,index=False,header=None)

        pd.set_option('display.max_colwidth', 500)   # FOR TABLE <th>
        

        html = new.to_html(escape=False,index=False,header=False).replace('&lt;','<').replace('&gt;', '>').replace(r'\n', '<br>').replace('table','table style="border-collapse: collapse"').replace('<td>','<td align="center">')
        # .replace('<td>','<td align="center">')
        soup = BeautifulSoup(html, features="lxml")
        table = soup.find('table')
        table_body = table.find('tbody')
        # td = table_body.find('td')
        
        rows = table_body.find_all('tr')
        data = []

        for row in rows:
            cols = row.find_all('td')
            # row.find('td')['colspan'] = '2'
           
            data.append(cols)

        if type=='Issue' and photo=="with photo" and status==None:
            data[0][0]['align'] = "left"
            data[0][1]['align'] = "left"

            data[1][0]['align'] = "left"
            data[1][1]['align'] = "left"

            data[2][0]['align'] = "left"
            data[2][1]['align'] = "left"

            data[3][0]['align'] = "left"
            data[3][1]['align'] = "left"

            data[4][0]['align'] = "left"
            data[4][1]['align'] = "left"

            data[5][0]['align'] = "left"
            data[5][1]['align'] = "left"

            data[6][0]['align'] = "left"
            data[6][1]['align'] = "left"

            data[7][0]['align'] = "left"
            data[7][1]['align'] = "left"
            
            data[4][1]['colspan'] = "2"
            data[4][2].decompose()

            data[3][1]['colspan'] = "5"
            data[3][2].decompose()
            data[3][3].decompose()
            data[3][4].decompose()
            data[3][5].decompose()

            data[8][1]['colspan'] = "2"
            data[8][2].decompose()

            data[8][0]['rowspan'] = "2"
            data[9][0].decompose()

            data[8][3]['rowspan'] = "2"
            data[9][3].decompose()

            data[8][4]['rowspan'] = "2"
            data[9][4].decompose()

            data[8][5]['rowspan'] = "2"
            data[9][5].decompose()

            data[8][6]['rowspan'] = "2"
            data[9][6].decompose()
            data[8][7]['rowspan'] = "2"
            data[9][7].decompose()
            data[8][8]['rowspan'] = "2"
            data[9][8].decompose()
            data[8][9]['rowspan'] = "2"
            data[9][9].decompose()
            data[8][10]['rowspan'] = "2"
            data[9][10].decompose()

        
        if type=='Issue' and photo=="without photo" and status==None:
           
            data[0][0]['align'] = "left"
            data[0][1]['align'] = "left"

            data[1][0]['align'] = "left"
            data[1][1]['align'] = "left"

            data[2][0]['align'] = "left"
            data[2][1]['align'] = "left"

            data[3][0]['align'] = "left"
            data[3][1]['align'] = "left"

            data[4][0]['align'] = "left"
            data[4][1]['align'] = "left"

            data[5][0]['align'] = "left"
            data[5][1]['align'] = "left"

            data[6][0]['align'] = "left"
            data[6][1]['align'] = "left"

            data[7][0]['align'] = "left"
            data[7][1]['align'] = "left"


            data[3][1]['colspan'] = "5"
            data[3][2].decompose()
            data[3][3].decompose()
            data[3][4].decompose()
            data[3][5].decompose()

            data[4][1]['colspan'] = "2"
            data[4][2].decompose()

            data[8][1]['colspan'] = "2"
            data[8][2].decompose()

            data[8][0]['rowspan'] = "2"
            data[9][0].decompose()

            data[8][3]['rowspan'] = "2"
            data[9][3].decompose()

            data[8][4]['rowspan'] = "2"
            data[9][4].decompose()

            data[8][5]['rowspan'] = "2"
            data[9][5].decompose()

            data[8][6]['rowspan'] = "2"
            data[9][6].decompose()
            data[8][7]['rowspan'] = "2"
            data[9][7].decompose()
            data[8][8]['rowspan'] = "2"
            data[9][8].decompose()
            data[8][9]['rowspan'] = "2"
            data[9][9].decompose()

        if type == 'Issue' and name=="Compliance" and photo=="with photo" and status:

            data[0][0]['align'] = "left"
            data[0][1]['align'] = "left"

            data[0][0]['width'] = "100"

            data[1][0]['align'] = "left"
            data[1][1]['align'] = "left"

            data[2][0]['align'] = "left"
            data[2][1]['align'] = "left"

            data[3][0]['align'] = "left"
            data[3][1]['align'] = "left"

            data[4][0]['align'] = "left"
            data[4][1]['align'] = "left"

            data[5][0]['align'] = "left"
            data[5][1]['align'] = "left"

            data[6][0]['align'] = "left"
            data[6][1]['align'] = "left"
            # data[6][0]['colspan'] = "2"
            # data[6][1].decompose()
            data[7][0]['align'] = "left"
            data[7][1]['align'] = "left"
            
            data[3][1]['colspan'] = "5"
            data[3][2].decompose()
            data[3][3].decompose()
            data[3][4].decompose()
            data[3][5].decompose()

            data[4][1]['colspan'] = "3"
            data[4][2].decompose()
            data[4][3].decompose()
            # data[4][4].decompose()
            # data[4][5].decompose()

            data[8][1]['colspan'] = "2"
            data[8][2].decompose()

            data[8][12]['colspan'] ="2"
            data[8][13].decompose()

            data[8][14]['colspan'] = "3"
            data[8][15].decompose()
            data[8][16].decompose()

            data[8][0]['rowspan'] = "2"
            data[9][0].decompose()

            data[8][3]['rowspan'] = "2"
            data[9][3].decompose()

            data[8][4]['rowspan'] = "2"
            data[9][4].decompose()

            data[8][5]['rowspan'] = "2"
            data[9][5].decompose()

            data[8][6]['rowspan'] = "2"
            data[9][6].decompose()

            data[8][7]['rowspan'] = "2"
            data[9][7].decompose()

            data[8][8]['rowspan'] = "2"
            data[9][8].decompose()

            data[8][9]['rowspan'] = "2"
            data[9][9].decompose()

            data[8][10]['rowspan'] = "2"
            data[9][10].decompose()

            data[8][11]['rowspan'] = "2"
            data[9][11].decompose()

            data[8][17]['rowspan'] = "2"
            data[9][17].decompose()

            data[8][18]['rowspan'] = "2"
            data[9][18].decompose()

            data[8][19]['rowspan'] = "2"
            data[9][19].decompose()

            data[8][20]['rowspan'] = "2"
            data[9][20].decompose()


        if type == 'Issue' and name=="Compliance" and photo=="without photo" and status:
            data[0][0]['width'] = "100"
            data[0][0]['align'] = "left"
            data[0][1]['align'] = "left"

            data[1][0]['align'] = "left"
            data[1][1]['align'] = "left"

            data[2][0]['align'] = "left"
            data[2][1]['align'] = "left"

            data[3][0]['align'] = "left"
            data[3][1]['align'] = "left"

            data[4][0]['align'] = "left"
            data[4][1]['align'] = "left"

            data[5][0]['align'] = "left"
            data[5][1]['align'] = "left"

            data[6][0]['align'] = "left"
            data[6][1]['align'] = "left"
            
            data[7][0]['align'] = "left"
            data[7][1]['align'] = "left"
            
            data[3][1]['colspan'] = "5"
            data[3][2].decompose()
            data[3][3].decompose()
            data[3][4].decompose()
            data[3][5].decompose()

            data[4][1]['colspan'] = "3"
            data[4][2].decompose()
            data[4][3].decompose()

            data[8][1]['colspan'] = "2"
            data[8][2].decompose()

            data[8][10]['colspan'] ="2"
            data[8][11].decompose()

            data[8][12]['colspan'] = "3"
            data[8][13].decompose()
            data[8][14].decompose()            

            data[8][0]['rowspan'] = "2"
            data[9][0].decompose()

            data[8][3]['rowspan'] = "2"
            data[9][3].decompose()

            data[8][4]['rowspan'] = "2"
            data[9][4].decompose()

            data[8][5]['rowspan'] = "2"
            data[9][5].decompose()

            data[8][6]['rowspan'] = "2"
            data[9][6].decompose()

            data[8][7]['rowspan'] = "2"
            data[9][7].decompose()

            data[8][8]['rowspan'] = "2"
            data[9][8].decompose()

            data[8][9]['rowspan'] = "2"
            data[9][9].decompose()

            data[8][15]['rowspan'] = "2"
            data[9][15].decompose()

            data[8][16]['rowspan'] = "2"
            data[9][16].decompose()

            data[8][17]['rowspan'] = "2"
            data[9][17].decompose()

            data[8][18]['rowspan'] = "2"
            data[9][18].decompose()


        if type=="Investigation" and photo=="with photo":
            data[0][0]['align'] = "left"
            data[0][1]['align'] = "left"

            data[1][0]['align'] = "left"
            data[1][1]['align'] = "left"

            data[2][0]['align'] = "left"
            data[2][1]['align'] = "left"

            data[3][0]['align'] = "left"
            data[3][1]['align'] = "left"

            data[4][0]['align'] = "left"
            data[4][1]['align'] = "left"

            data[5][0]['align'] = "left"
            data[5][1]['align'] = "left"

            data[6][0]['align'] = "left"
            data[6][1]['align'] = "left"

            data[7][0]['align'] = "left"
            data[7][1]['align'] = "left"

            data[3][1]['colspan'] = "5"
            data[3][2].decompose()
            data[3][3].decompose()
            data[3][4].decompose()
            data[3][5].decompose()

            data[8][2]['colspan'] ="2"
            data[8][3].decompose()

            data[8][0]['rowspan'] = "2"
            data[9][0].decompose()

            data[8][1]['rowspan'] = "2"
            data[9][1].decompose()

            data[8][4]['rowspan'] = "2"
            data[9][4].decompose()

            data[8][5]['rowspan'] = "2"
            data[9][5].decompose()

            data[8][6]['rowspan'] = "2"
            data[9][6].decompose()

            data[8][7]['rowspan'] = "2"
            data[9][7].decompose()

            data[8][8]['rowspan'] = "2"
            data[9][8].decompose()

            data[8][9]['rowspan'] = "2"
            data[9][9].decompose()

            data[8][10]['rowspan'] = "2"
            data[9][10].decompose()

            data[8][11]['rowspan'] = "2"
            data[9][11].decompose()

            # data[8][12]['rowspan'] = "2"
            # data[9][12].decompose()

        if type=="Investigation" and photo=="without photo":
            data[0][0]['align'] = "left"
            data[0][1]['align'] = "left"

            data[1][0]['align'] = "left"
            data[1][1]['align'] = "left"

            data[2][0]['align'] = "left"
            data[2][1]['align'] = "left"

            data[3][0]['align'] = "left"
            data[3][1]['align'] = "left"

            data[4][0]['align'] = "left"
            data[4][1]['align'] = "left"

            data[5][0]['align'] = "left"
            data[5][1]['align'] = "left"

            data[6][0]['align'] = "left"
            data[6][1]['align'] = "left"

            data[7][0]['align'] = "left"
            data[7][1]['align'] = "left"

            data[3][1]['colspan'] = "5"
            data[3][2].decompose()
            data[3][3].decompose()
            data[3][4].decompose()
            data[3][5].decompose()

            data[8][2]['colspan'] ="2"
            data[8][3].decompose()

            data[8][0]['rowspan'] = "2"
            data[9][0].decompose()

            data[8][1]['rowspan'] = "2"
            data[9][1].decompose()

            data[8][4]['rowspan'] = "2"
            data[9][4].decompose()

            data[8][5]['rowspan'] = "2"
            data[9][5].decompose()

            data[8][6]['rowspan'] = "2"
            data[9][6].decompose()

            data[8][7]['rowspan'] = "2"
            data[9][7].decompose()

            data[8][8]['rowspan'] = "2"
            data[9][8].decompose()

            data[8][9]['rowspan'] = "2"
            data[9][9].decompose()

            data[8][10]['rowspan'] = "2"
            data[9][10].decompose()

            data[8][11]['rowspan'] = "2"
            data[9][11].decompose()

        if type=="Accident" and accident_report_type== "Detail":
            data[0][0]['align'] = "left"
            data[0][1]['align'] = "left"

            data[1][0]['align'] = "left"
            data[1][1]['align'] = "left"

            data[2][0]['align'] = "left"
            data[2][1]['align'] = "left"

            data[3][0]['align'] = "left"
            data[3][1]['align'] = "left"

            data[4][0]['align'] = "left"
            data[4][1]['align'] = "left"

            data[5][0]['align'] = "left"
            data[5][1]['align'] = "left"

            # data[6][0]['align'] = "left"
            # data[6][1]['align'] = "left"

            # data[7][0]['align'] = "left"
            # data[7][1]['align'] = "left"

            data[3][1]['colspan'] = "3"
            data[3][2].decompose()
            data[3][3].decompose()
            # data[3][4].decompose()
            # data[3][5].decompose()

            data[7][0]['colspan'] = "18"
            data[7][1].decompose()
            data[7][2].decompose()
            data[7][3].decompose()
            data[7][4].decompose()
            data[7][5].decompose()
            data[7][6].decompose()
            data[7][7].decompose()
            data[7][8].decompose()
            data[7][9].decompose()
            data[7][10].decompose()
            data[7][11].decompose()
            data[7][12].decompose()
            data[7][13].decompose()
            data[7][14].decompose()
            data[7][15].decompose()
            data[7][16].decompose()
            data[7][17].decompose()

            data[8][0]['colspan'] = "3"
            data[8][1].decompose()
            data[8][2].decompose()
            
            data[8][12]['colspan'] = "4"
            data[8][13].decompose()
            data[8][14].decompose()

            data[-9][0]['align'] = "left"
            data[-9][0]['align'] = "left"
            data[-9][0]['colspan'] = "2"
            data[-9][1].decompose()

            data[-8][0]['align'] = "left"
            data[-8][0]['colspan'] ="10"
            data[-8][1].decompose()
            data[-8][2].decompose()
            data[-8][3].decompose()
            data[-8][4].decompose()
            data[-8][5].decompose()
            data[-8][6].decompose()
            data[-8][7].decompose()
            data[-8][8].decompose()
            data[-8][9].decompose()
            # data[27][10].decompose()
            # data[27][11].decompose()

            data[-7][0]['align'] = "left"
            data[-7][0]['colspan'] ="10"
            data[-7][1].decompose()
            data[-7][2].decompose()
            data[-7][3].decompose()
            data[-7][4].decompose()
            data[-7][5].decompose()
            data[-7][6].decompose()
            data[-7][7].decompose()
            data[-7][8].decompose()
            data[-7][9].decompose()
            # data[28][10].decompose()
            # data[28][11].decompose()

            data[-6][0]['align'] = "left"
            data[-6][0]['colspan'] ="13"
            data[-6][1].decompose()
            data[-6][2].decompose()
            data[-6][3].decompose()
            data[-6][4].decompose()
            data[-6][5].decompose()
            data[-6][6].decompose()
            data[-6][7].decompose()
            data[-6][8].decompose()
            data[-6][9].decompose()
            data[-6][10].decompose()
            data[-6][11].decompose()
            data[-6][12].decompose()

            data[-5][0]['align'] = "left"
            data[-5][0]['colspan'] ="10"
            data[-5][1].decompose()
            data[-5][2].decompose()
            data[-5][3].decompose()
            data[-5][4].decompose()
            data[-5][5].decompose()
            data[-5][6].decompose()
            data[-5][7].decompose()
            data[-5][8].decompose()
            data[-5][9].decompose()
            # data[30][10].decompose()

            data[-4][0]['align'] = "left"
            data[-4][0]['colspan'] ="10"
            data[-4][1].decompose()
            data[-4][2].decompose()
            data[-4][3].decompose()
            data[-4][4].decompose()
            data[-4][5].decompose()
            data[-4][6].decompose()
            data[-4][7].decompose()
            data[-4][8].decompose()
            data[-4][9].decompose()
            # data[31][10].decompose()

            data[-3][0]['align'] = "left"
            data[-3][0]['colspan'] ="10"
            data[-3][1].decompose()
            data[-3][2].decompose()
            data[-3][3].decompose()
            data[-3][4].decompose()
            data[-3][5].decompose()
            data[-3][6].decompose()
            data[-3][7].decompose()
            data[-3][8].decompose()
            data[-3][9].decompose()
            # data[32][10].decompose()

            data[-2][0]['align'] = "left"
            data[-2][0]['colspan'] ="10"
            data[-2][1].decompose()
            data[-2][2].decompose()
            data[-2][3].decompose()
            data[-2][4].decompose()
            data[-2][5].decompose()
            data[-2][6].decompose()
            data[-2][7].decompose()
            data[-2][8].decompose()
            data[-2][9].decompose()
            # data[-2][10].decompose()

            data[-1][0]['align'] = "left"
            data[-1][0]['colspan'] ="10"
            data[-1][1].decompose()
            data[-1][2].decompose()
            data[-1][3].decompose()
            data[-1][4].decompose()
            data[-1][5].decompose()
            data[-1][6].decompose()
            data[-1][7].decompose()
            data[-1][8].decompose()
            data[-1][9].decompose()
  
        if type=="Accident" and accident_report_type=="Summary":
            data[0][0]['align'] = "left"
            data[0][1]['align'] = "left"

            data[1][0]['align'] = "left"
            data[1][1]['align'] = "left"

            data[2][0]['align'] = "left"
            data[2][1]['align'] = "left"

            data[3][0]['align'] = "left"
            data[3][1]['align'] = "left"

            data[4][0]['align'] = "left"
            data[4][1]['align'] = "left"

            data[5][0]['align'] = "left"
            data[5][1]['align'] = "left"

            # data[6][0]['align'] = "left"
            # data[6][1]['align'] = "left"

            # data[7][0]['align'] = "left"
            # data[7][1]['align'] = "left"

            data[3][1]['colspan'] = "5"
            data[3][2].decompose()
            data[3][3].decompose()
            data[3][4].decompose()
            data[3][5].decompose()

            data[7][2]['colspan'] = "2"
            data[7][3].decompose()
            data[7][4]['colspan'] = "2"
            data[7][5].decompose()
            data[7][6]['colspan'] = "2"
            data[7][7].decompose()

            data[7][0]['rowspan'] = "2"
            data[8][0].decompose()

            data[7][1]['rowspan'] = "2"
            data[8][1].decompose()


        h = str(soup)

        pisa_context = pisa.CreatePDF(h)
        response = pisa_context.dest.getvalue()
        return h
        # return response
    
    return wb    
    